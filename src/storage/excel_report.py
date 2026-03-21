"""
GovScheme SuperAgent â€” Excel Report Generator
Produces multi-sheet Excel workbooks for tracking, dashboarding, and daily reports.
Sheets:
  1. Master Tracker      â€” All schemes with dates, fees, status, amounts
  2. New Schemes          â€” Schemes discovered in this run
  3. Approaching Deadlinesâ€” Schemes with deadlines within 7/30 days
  4. Sector Summary       â€” Pivot by sector with counts
  5. State Summary        â€” Pivot by state with counts
  6. Status Summary       â€” Active/Expired/Upcoming/Closed breakdown
  7. Daily Run History    â€” Historical log of all daily crawl runs
  8. Change Log           â€” All detected changes across runs
"""
from __future__ import annotations
import logging
from datetime import date, datetime
from pathlib import Path
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from src.agents.models import DailyRunReport
from src.storage.database import SchemeDatabase
logger = logging.getLogger("excel_report")
# â”€â”€ Additional Style Constants for Exam Sheets â”€â”€
UPCOMING_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")  # Blue
ORANGE_FILL = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="E8D0F0", end_color="E8D0F0", fill_type="solid")
TEAL_FILL = PatternFill(start_color="B2DFDB", end_color="B2DFDB", fill_type="solid")
GREY_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
# â”€â”€ Style Constants â”€â”€
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DATE_FONT = Font(name="Arial", size=10, color="0066CC")
FEE_FONT = Font(name="Arial", size=10, color="CC6600")
NEW_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
DEADLINE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
EXPIRED_FILL = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
ACTIVE_FILL = PatternFill(start_color="DFF0D8", end_color="DFF0D8", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color="1F4E79"),
    right=Side(style="thin", color="1F4E79"),
    top=Side(style="thin", color="1F4E79"),
    bottom=Side(style="medium", color="1F4E79"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)
class ExcelReportGenerator:
    """Generates multi-sheet Excel reports for scheme + exam tracking."""
    def __init__(self, db: SchemeDatabase, output_dir: str = "./output", exam_db=None):
        self.db = db
        self.exam_db = exam_db  # ExamDatabase instance (optional, V3)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    def generate_full_report(
        self,
        daily_report: Optional[DailyRunReport] = None,
        exam_report=None,
    ) -> str:
        """Generate the comprehensive Excel report. Returns the file path."""
        wb = Workbook()
        # Remove the default sheet
        wb.remove(wb.active)
        # â”€â”€ Sheet 1: Master Tracker â”€â”€
        self._build_master_tracker(wb)
        # â”€â”€ Sheet 2: New Schemes (this run) â”€â”€
        if daily_report:
            self._build_new_schemes_sheet(wb, daily_report.run_date)
        # â”€â”€ Sheet 3: Approaching Deadlines â”€â”€
        self._build_deadlines_sheet(wb)
        # â”€â”€ Sheet 4: Sector Summary â”€â”€
        self._build_sector_summary(wb)
        # â”€â”€ Sheet 5: State Summary â”€â”€
        self._build_state_summary(wb)
        # â”€â”€ Sheet 6: Status Summary â”€â”€
        self._build_status_summary(wb)
        # â”€â”€ Sheet 7: Daily Run History â”€â”€
        self._build_run_history(wb)
        # â”€â”€ Sheet 8: Change Log â”€â”€
        if daily_report:
            self._build_change_log(wb, daily_report.run_id)
        # â”€â”€ Sheet 9: Dashboard (charts) â”€â”€
        self._build_dashboard_sheet(wb)
        # â•â•â• V3: EXAM SHEETS (10â€“15) â•â•â•
        if self.exam_db:
            self._build_exam_master_tracker(wb)        # Sheet 10
            self._build_exam_application_open(wb)      # Sheet 11
            self._build_exam_calendar(wb)              # Sheet 12
            self._build_exam_deadlines(wb)             # Sheet 13
            self._build_exam_category_summary(wb)      # Sheet 14
            self._build_exam_state_summary(wb)         # Sheet 15
        # Save
        today = date.today().isoformat()
        filename = f"GovScheme_Report_{today}.xlsx"
        filepath = self.output_dir / "reports" / filename
        filepath.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(filepath))
        logger.info("Excel report generated: %s (%d sheets)", filepath, len(wb.sheetnames))
        return str(filepath)
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # Sheet 1: Master Tracker
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    def _build_master_tracker(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Master Tracker")
        headers = [
            "S.No", "Scheme Name", "Level", "State/UT", "Sector", "Type",
            "Status", "Start Date", "End Date", "Application Start",
            "Application Deadline", "Application Fee", "Min Amount",
            "Max Amount", "Frequency", "Target Group",
            "Age Limit", "Income Limit", "Gender", "Caste/Category",
            "Ministry/Dept", "Official Website", "Helpline",
            "Summary", "Eligibility", "Documents Required",
            "Source Portal", "Source URL", "First Seen", "Last Seen",
            "Times Seen", "Confidence", "Folder Path",
        ]
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = HEADER_BORDER
        # Freeze top row
        ws.freeze_panes = "A2"
        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        # Write data
        schemes = self.db.get_all_schemes()
        for idx, s in enumerate(schemes, 1):
            row = idx + 1
            docs = s.get("documents_required", "[]")
            try:
                import json
                docs_list = json.loads(docs) if docs else []
                docs_str = "; ".join(docs_list) if isinstance(docs_list, list) else str(docs)
            except Exception:
                docs_str = str(docs)
            values = [
                idx, s.get("clean_name"), s.get("level"), s.get("state"),
                s.get("sector"), s.get("scheme_type"),
                s.get("scheme_status", "Unknown"),
                s.get("start_date"), s.get("end_date"),
                s.get("application_start_date"), s.get("application_end_date") or s.get("application_deadline"),
                s.get("application_fee"), s.get("fund_amount_min"),
                s.get("fund_amount_max"), s.get("frequency"),
                s.get("target_group"), s.get("age_limit"),
                s.get("income_limit"), s.get("gender_eligibility"),
                s.get("caste_eligibility"),
                s.get("nodal_ministry") or s.get("nodal_department"),
                s.get("official_website"), s.get("helpline"),
                s.get("summary"), s.get("eligibility"),
                docs_str,
                s.get("source_portal"), s.get("source_url"),
                s.get("first_seen_date"), s.get("last_seen_date"),
                s.get("times_seen"), s.get("classification_confidence"),
                s.get("folder_path"),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = THIN_BORDER
                cell.alignment = WRAP_ALIGN
                # Conditional formatting
                if col == 7:  # Status
                    if val == "Active":
                        cell.fill = ACTIVE_FILL
                    elif val == "Expired":
                        cell.fill = EXPIRED_FILL
                    elif val == "Closed":
                        cell.fill = DEADLINE_FILL
                if col in (8, 9, 10, 11):  # Date columns
                    cell.font = DATE_FONT
                if col in (12, 13, 14):  # Fee/Amount columns
                    cell.font = FEE_FONT
        # Set column widths
        col_widths = [
            6, 40, 10, 18, 20, 14, 10, 14, 14, 14, 14, 14, 14, 14, 12,
            20, 12, 18, 10, 16, 25, 30, 18, 50, 40, 40, 16, 35, 12, 12,
            8, 8, 30,
        ]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # Sheet 2: New Schemes
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    def _build_new_schemes_sheet(self, wb: Workbook, since_date: str) -> None:
        ws = wb.create_sheet("New Schemes")
        headers = [
            "S.No", "Scheme Name", "Level", "State/UT", "Sector", "Type",
            "Start Date", "End Date", "Application Fee",
            "Fund Amount", "Target Group", "Summary", "Source URL",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            cell.alignment = CENTER_ALIGN
            cell.border = HEADER_BORDER
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        new_schemes = self.db.get_new_since(since_date)
        for idx, s in enumerate(new_schemes, 1):
            row = idx + 1
            amount = s.get("fund_amount_min") or ""
            if s.get("fund_amount_max"):
                amount = f"{amount} - {s['fund_amount_max']}" if amount else s["fund_amount_max"]
            values = [
                idx, s.get("clean_name"), s.get("level"), s.get("state"),
                s.get("sector"), s.get("scheme_type"),
                s.get("start_date"), s.get("end_date"),
                s.get("application_fee"), amount,
                s.get("target_group"), s.get("summary"), s.get("source_url"),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = NEW_FILL
                cell.border = THIN_BORDER
                cell.alignment = WRAP_ALIGN
        widths = [6, 40, 10, 18, 20, 14, 14, 14, 14, 20, 20, 50, 35]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # Sheet 3: Approaching Deadlines
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    def _build_deadlines_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Approaching Deadlines")
        headers = [
            "S.No", "Scheme Name", "Level", "State/UT", "Sector",
            "Application Deadline", "End Date", "Days Remaining",
            "Application Fee", "Fund Amount", "Target Group",
            "How to Apply", "Source URL",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
            cell.alignment = CENTER_ALIGN
            cell.border = HEADER_BORDER
        ws.freeze_panes = "A2"
        # Get deadlines within 30 days
        deadlines_7 = self.db.get_approaching_deadlines(7)
        deadlines_30 = self.db.get_approaching_deadlines(30)
        # Combine and deduplicate
        seen_ids = set()
        all_deadlines = []
        for s in deadlines_7 + deadlines_30:
            sid = s.get("scheme_id")
            if sid not in seen_ids:
                seen_ids.add(sid)
                all_deadlines.append(s)
        for idx, s in enumerate(all_deadlines, 1):
            row = idx + 1
            deadline = s.get("application_end_date") or s.get("end_date") or ""
            days_remaining = ""
            try:
                if deadline:
                    dl = date.fromisoformat(deadline[:10])
                    days_remaining = (dl - date.today()).days
            except (ValueError, TypeError):
                pass
            amount = s.get("fund_amount_min") or ""
            if s.get("fund_amount_max"):
                amount = f"{amount} - {s['fund_amount_max']}" if amount else s["fund_amount_max"]
            values = [
                idx, s.get("clean_name"), s.get("level"), s.get("state"),
                s.get("sector"), s.get("application_end_date"),
                s.get("end_date"), days_remaining,
                s.get("application_fee"), amount,
                s.get("target_group"), "", s.get("source_url"),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = THIN_BORDER
                cell.alignment = WRAP_ALIGN
                if col == 8 and isinstance(days_remaining, int):
                    if days_remaining <= 3:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
                    elif days_remaining <= 7:
                        cell.fill = DEADLINE_FILL
                    elif days_remaining <= 30:
                        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        widths = [6, 40, 10, 18, 20, 16, 16, 14, 14, 20, 20, 30, 35]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # Sheet 4: Sector Summary
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    def _build_sector_summary(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Sector Summary")
        stats = self.db.get_stats()
        headers = ["Sector", "Count", "% of Total"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = HEADER_BORDER
        total = stats["total"] or 1
        for idx, (sector, count) in enumerate(
            sorted(stats["by_sector"].items(), key=lambda x: -x[1]), 1
        ):
            row = idx + 1
            ws.cell(row=row, column=1, value=sector.replace("_", " ")).border = THIN_BORDER
            ws.cell(row=row, column=2, value=count).border = THIN_BORDER
            pct_cell = ws.cell(row=row, column=3, value=count / total)
            pct_cell.number_format = "0.0%"
            pct_cell.border = THIN_BORDER
        # Total row
        total_row = len(stats["by_sector"]) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row-1})")
        ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row-1})")
        ws.cell(row=total_row, column=3).number_format = "0.0%"
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # Sheet 5: State Summary
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    def _build_state_summary(self, wb: Workbook) -> None:
        ws = wb.create_sheet("State Summary")
        stats = self.db.get_stats()
        headers = ["State/UT", "Count", "% of Total"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = HEADER_BORDER
        total_state = sum(stats["by_state"].values()) or 1
        for idx, (state, count) in enumerate(
            sorted(stats["by_state"].items(), key=lambda x: -x[1]), 1
        ):
            row = idx + 1
            ws.cell(row=row, column=1, value=state.replace("_", " ")).border = THIN_BORDER
            ws.cell(row=row, column=2, value=count).border = THIN_BORDER
            pct_cell = ws.cell(row=row, column=3, value=count / total_state)
            pct_cell.number_format = "0.0%"
            pct_cell.border = THIN_BORDER
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # Sheet 6: Status Summary
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    def _build_status_summary(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Status Summary")
        stats = self.db.get_stats()
        headers = ["Status", "Count", "% of Total"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
        total = stats["total"] or 1
        status_fills = {
            "Active": ACTIVE_FILL,
            "Expired": EXPIRED_FILL,
            "Closed": DEADLINE_FILL,
            "Upcoming": PatternFill(start_color="D9EDF7", end_color="D9EDF7", fill_type="solid"),
            "Unknown": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
        }
        for idx, (status, count) in enumerate(
            sorted(stats["by_status"].items(), key=lambda x: -x[1]), 1
        ):
            row = idx + 1
            cell_name = ws.cell(row=row, column=1, value=status)
            cell_name.fill = status_fills.get(status, PatternFill())
            cell_name.border = THIN_BORDER
            ws.cell(row=row, column=2, value=count).border = THIN_BORDER
            pct = ws.cell(row=row, column=3, value=count / total)
            pct.number_format = "0.0%"
            pct.border = THIN_BORDER
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # Sheet 7: Daily Run History
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    def _build_run_history(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Run History")
        headers = [
            "Run Date", "Started At", "Completed At", "Total in DB",
            "New", "Updated", "Closed", "Unchanged", "Errors",
            "Duration (sec)", "Excel Report",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = HEADER_BORDER
        ws.freeze_panes = "A2"
        runs = self.db.get_run_history(60)
        for idx, r in enumerate(runs, 1):
            row = idx + 1
            values = [
                r.get("run_date"), r.get("started_at"), r.get("completed_at"),
                r.get("total_in_db"), r.get("new_schemes"),
                r.get("updated_schemes"), r.get("closed_schemes"),
                r.get("unchanged_schemes"), r.get("errors"),
                round(r.get("elapsed_seconds", 0), 1),
                r.get("excel_report_path"),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = THIN_BORDER
                if col == 5 and val and val > 0:  # New schemes
                    cell.fill = NEW_FILL
                    cell.font = Font(bold=True, color="006600")
        widths = [14, 22, 22, 12, 8, 10, 10, 12, 8, 12, 35]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # Sheet 8: Change Log
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    def _build_change_log(self, wb: Workbook, run_id: str) -> None:
        ws = wb.create_sheet("Change Log")
        headers = [
            "Scheme Name", "Level", "Sector", "Change Type",
            "Field Changed", "Old Value", "New Value", "Detected At",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="F57C00", end_color="F57C00", fill_type="solid")
            cell.alignment = CENTER_ALIGN
        changes = self.db.get_changes_for_run(run_id)
        for idx, c in enumerate(changes, 1):
            row = idx + 1
            values = [
                c.get("clean_name"), c.get("level"), c.get("sector"),
                c.get("change_type"), c.get("field_changed"),
                c.get("old_value", "")[:100], c.get("new_value", "")[:100],
                c.get("detected_at"),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = THIN_BORDER
        widths = [40, 10, 20, 14, 18, 30, 30, 22]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # Sheet 9: Dashboard (Charts)
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    def _build_dashboard_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Dashboard")
        stats = self.db.get_stats()
        # Title
        ws.merge_cells("A1:F1")
        title_cell = ws.cell(row=1, column=1, value="GovScheme SuperAgent â€” Dashboard")
        title_cell.font = Font(name="Arial", size=16, bold=True, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=1, value=f"Report Date: {date.today().isoformat()}")
        ws.cell(row=2, column=1).font = Font(italic=True, color="666666")
        # KPI Cards (row 4)
        kpis = [
            ("Total Schemes", stats["total"]),
            ("Active", stats["active"]),
            ("Expired", stats.get("by_status", {}).get("Expired", 0)),
            ("Upcoming", stats.get("by_status", {}).get("Upcoming", 0)),
        ]
        for i, (label, val) in enumerate(kpis):
            col = (i * 2) + 1
            ws.cell(row=4, column=col, value=label).font = Font(size=9, color="666666")
            num_cell = ws.cell(row=5, column=col, value=val)
            num_cell.font = Font(size=20, bold=True, color="1F4E79")
        # Sector data for chart (row 8+)
        ws.cell(row=8, column=1, value="Sector").font = Font(bold=True)
        ws.cell(row=8, column=2, value="Count").font = Font(bold=True)
        sorted_sectors = sorted(stats["by_sector"].items(), key=lambda x: -x[1])
        for idx, (sector, count) in enumerate(sorted_sectors[:15], 1):
            ws.cell(row=8 + idx, column=1, value=sector.replace("_", " "))
            ws.cell(row=8 + idx, column=2, value=count)
        # Bar chart for sectors
        if sorted_sectors:
            chart = BarChart()
            chart.type = "bar"
            chart.title = "Schemes by Sector"
            chart.style = 10
            chart.y_axis.title = "Number of Schemes"
            chart.x_axis.title = "Sector"
            data_ref = Reference(ws, min_col=2, min_row=8, max_row=8 + min(15, len(sorted_sectors)))
            cats_ref = Reference(ws, min_col=1, min_row=9, max_row=8 + min(15, len(sorted_sectors)))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.shape = 4
            chart.width = 30
            chart.height = 15
            ws.add_chart(chart, "D8")
        # Level data for pie chart
        level_start = 8 + len(sorted_sectors) + 3
        ws.cell(row=level_start, column=1, value="Level").font = Font(bold=True)
        ws.cell(row=level_start, column=2, value="Count").font = Font(bold=True)
        for idx, (level, count) in enumerate(stats["by_level"].items(), 1):
            ws.cell(row=level_start + idx, column=1, value=level)
            ws.cell(row=level_start + idx, column=2, value=count)
        if stats["by_level"]:
            pie = PieChart()
            pie.title = "Schemes by Level"
            pie.style = 10
            data_ref = Reference(ws, min_col=2, min_row=level_start, max_row=level_start + len(stats["by_level"]))
            cats_ref = Reference(ws, min_col=1, min_row=level_start + 1, max_row=level_start + len(stats["by_level"]))
            pie.add_data(data_ref, titles_from_data=True)
            pie.set_categories(cats_ref)
            pie.width = 16
            pie.height = 12
            ws.add_chart(pie, "D" + str(level_start))
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 12
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # EXAM SHEETS (V3 â€” Sheets 10â€“15)
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    def _apply_exam_header(self, ws, headers: list[str]) -> None:
        """Apply header formatting to an exam sheet."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = HEADER_BORDER
    def _apply_exam_status_fill(self, cell, status: str) -> None:
        """Color-code exam status cells."""
        fills = {
            "Application_Open": ACTIVE_FILL,
            "Upcoming": UPCOMING_FILL,
            "Application_Closed": ORANGE_FILL,
            "Admit_Card_Out": ORANGE_FILL,
            "Exam_Ongoing": PURPLE_FILL,
            "Result_Awaited": TEAL_FILL,
            "Completed": GREY_FILL,
        }
        fill = fills.get(status)
        if fill:
            cell.fill = fill
    def _apply_deadline_urgency(self, cell, days) -> None:
        """Apply urgency colors: â‰¤3d red bold, â‰¤7d red, â‰¤30d yellow."""
        if days is None:
            return
        try:
            d = int(days)
        except (ValueError, TypeError):
            return
        if d <= 3:
            cell.fill = DEADLINE_FILL
            cell.font = Font(name="Arial", size=10, bold=True, color="CC0000")
        elif d <= 7:
            cell.fill = DEADLINE_FILL
        elif d <= 30:
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    def _build_exam_master_tracker(self, wb: Workbook) -> None:
        """Sheet 10: Exam Master Tracker â€” all exams with 40 columns."""
        ws = wb.create_sheet("Exam Master Tracker")
        headers = [
            "S.No", "Exam Name", "Short Name", "Conducting Body", "Category",
            "Level", "State", "Exam Cycle", "Status",
            "Notification Date", "Application Start", "Application End âš ï¸",
            "Fee Payment Deadline", "Correction Window",
            "Prelims Date", "Mains Date", "Admit Card Date",
            "Result Date", "Final Result Date", "Joining Date",
            "Fee (Gen) â‚¹", "Fee (OBC) â‚¹", "Fee (SC/ST) â‚¹",
            "Fee (Female) â‚¹", "Fee (EWS) â‚¹", "Fee (PwD) â‚¹", "Is Free?",
            "Total Vacancies", "Age Min", "Age Max",
            "Qualification", "Physical Standards",
            "Domicile Required", "Gender Restriction",
            "Apply Online URL", "Notification PDF", "Syllabus URL", "Official Website",
            "Days Until App Close", "Days Until Exam",
        ]
        self._apply_exam_header(ws, headers)
        exams = self.exam_db.get_all_exams()
        import json as _json
        for idx, exam in enumerate(exams, 1):
            row = idx + 1
            # Parse phases JSON to get prelims/mains dates
            phases = []
            try:
                phases = _json.loads(exam.get("phases_json") or "[]")
            except (ValueError, TypeError):
                pass
            prelims_date = mains_date = admit_card = None
            for p in phases:
                name = (p.get("phase_name") or "").lower()
                if "prelim" in name or "cbt" in name or "written" in name:
                    prelims_date = prelims_date or p.get("exam_date_start")
                    admit_card = admit_card or p.get("admit_card_date")
                elif "main" in name or "phase 2" in name or "interview" in name:
                    mains_date = mains_date or p.get("exam_date_start")
            # Compute days
            today = date.today()
            days_app_close = None
            days_exam = None
            app_end = exam.get("application_end_date")
            if app_end:
                try:
                    d = date.fromisoformat(app_end)
                    if d >= today:
                        days_app_close = (d - today).days
                except ValueError:
                    pass
            if prelims_date:
                try:
                    d = date.fromisoformat(prelims_date)
                    if d >= today:
                        days_exam = (d - today).days
                except ValueError:
                    pass
            values = [
                idx, exam.get("clean_exam_name") or exam.get("exam_name"), exam.get("short_name"),
                exam.get("conducting_body"), exam.get("exam_category"),
                exam.get("exam_level"), exam.get("state"), exam.get("exam_cycle"),
                exam.get("exam_status"),
                exam.get("notification_date"), exam.get("application_start_date"),
                exam.get("application_end_date"), exam.get("fee_payment_deadline"),
                f"{exam.get('correction_window_start', '') or ''} - {exam.get('correction_window_end', '') or ''}".strip(" - ") or None,
                prelims_date, mains_date, admit_card,
                exam.get("result_date"), exam.get("final_result_date"), exam.get("joining_date"),
                exam.get("fee_general"), exam.get("fee_obc"), exam.get("fee_sc_st"),
                exam.get("fee_female"), exam.get("fee_ews"), exam.get("fee_pwd"),
                "Yes" if exam.get("is_free") else "No",
                exam.get("total_vacancies"), exam.get("age_min"), exam.get("age_max"),
                exam.get("qualification"), exam.get("physical_standards"),
                exam.get("domicile_required"), exam.get("gender_restriction"),
                exam.get("apply_online_url"), exam.get("official_notification_url"),
                exam.get("syllabus_url"), exam.get("official_website"),
                days_app_close, days_exam,
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = THIN_BORDER
                cell.alignment = WRAP_ALIGN
            # Status coloring (col 9)
            self._apply_exam_status_fill(ws.cell(row=row, column=9), exam.get("exam_status", ""))
            # Application End urgency (col 12)
            self._apply_deadline_urgency(ws.cell(row=row, column=12), days_app_close)
            # Days Until App Close (col 39)
            self._apply_deadline_urgency(ws.cell(row=row, column=39), days_app_close)
            # Days Until Exam (col 40)
            if days_exam is not None and days_exam <= 7:
                ws.cell(row=row, column=40).fill = ORANGE_FILL
            elif days_exam is not None and days_exam <= 30:
                ws.cell(row=row, column=40).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            # Is Free coloring (col 27)
            if exam.get("is_free"):
                ws.cell(row=row, column=27).fill = ACTIVE_FILL
            # Hyperlinks for URL columns
            for url_col in [35, 36, 37, 38]:
                url_val = ws.cell(row=row, column=url_col).value
                if url_val and isinstance(url_val, str) and url_val.startswith("http"):
                    ws.cell(row=row, column=url_col).hyperlink = url_val
                    ws.cell(row=row, column=url_col).font = Font(color="0563C1", underline="single")
        # Auto-filter
        if exams:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(exams) + 1}"
        # Freeze panes
        ws.freeze_panes = "C2"
        # Column widths
        widths = {"A": 6, "B": 40, "C": 12, "D": 18, "E": 16, "F": 8, "G": 15, "H": 10, "I": 18,
                  "J": 12, "K": 12, "L": 14, "M": 12, "N": 18, "O": 12, "P": 12, "Q": 12,
                  "R": 12, "S": 12, "T": 12}
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w
    def _build_exam_application_open(self, wb: Workbook) -> None:
        """Sheet 11: Application Open Now â€” exams currently accepting applications."""
        ws = wb.create_sheet("Application Open Now")
        headers = [
            "S.No", "Exam Name", "Conducting Body", "Category",
            "Application Start", "Application End âš ï¸", "Fee (Gen) â‚¹",
            "Total Vacancies", "Days Left", "Apply URL",
        ]
        self._apply_exam_header(ws, headers)
        open_exams = self.exam_db.get_application_open()
        today = date.today()
        for idx, exam in enumerate(open_exams, 1):
            row = idx + 1
            days_left = None
            app_end = exam.get("application_end_date")
            if app_end:
                try:
                    days_left = (date.fromisoformat(app_end) - today).days
                except ValueError:
                    pass
            values = [
                idx, exam.get("clean_exam_name") or exam.get("exam_name"),
                exam.get("conducting_body"), exam.get("exam_category"),
                exam.get("application_start_date"), exam.get("application_end_date"),
                exam.get("fee_general"), exam.get("total_vacancies"),
                days_left, exam.get("apply_online_url"),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = THIN_BORDER
                cell.alignment = WRAP_ALIGN
            # Urgency on Days Left (col 9)
            self._apply_deadline_urgency(ws.cell(row=row, column=9), days_left)
            # Apply URL hyperlink (col 10)
            url_val = ws.cell(row=row, column=10).value
            if url_val and isinstance(url_val, str) and url_val.startswith("http"):
                ws.cell(row=row, column=10).hyperlink = url_val
                ws.cell(row=row, column=10).font = Font(color="0563C1", underline="single")
        ws.freeze_panes = "C2"
        widths = {"A": 6, "B": 40, "C": 20, "D": 16, "E": 14, "F": 14, "G": 10, "H": 12, "I": 10, "J": 35}
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w
    def _build_exam_calendar(self, wb: Workbook) -> None:
        """Sheet 12: Exam Calendar â€” all upcoming events sorted chronologically."""
        ws = wb.create_sheet("Exam Calendar")
        headers = ["Date", "Exam Name", "Event Type", "Conducting Body", "Category", "Apply URL", "Days Away"]
        self._apply_exam_header(ws, headers)
        import json as _json
        all_exams = self.exam_db.get_all_exams()
        today = date.today()
        events: list[dict] = []
        for exam in all_exams:
            name = exam.get("clean_exam_name") or exam.get("exam_name", "")
            body = exam.get("conducting_body", "")
            cat = exam.get("exam_category", "")
            apply_url = exam.get("apply_online_url", "")
            for field, event_type in [
                ("application_start_date", "Application Open"),
                ("application_end_date", "Application Close"),
                ("result_date", "Result"),
                ("final_result_date", "Final Result"),
            ]:
                dt = exam.get(field)
                if dt:
                    events.append({"date": dt, "exam": name, "type": event_type,
                                    "body": body, "cat": cat, "url": apply_url})
            phases_raw = exam.get("phases_json")
            if phases_raw:
                try:
                    for phase in _json.loads(phases_raw):
                        pn = phase.get("phase_name", "")
                        if phase.get("admit_card_date"):
                            events.append({"date": phase["admit_card_date"], "exam": name,
                                            "type": f"Admit Card ({pn})", "body": body, "cat": cat, "url": apply_url})
                        if phase.get("exam_date_start"):
                            events.append({"date": phase["exam_date_start"], "exam": name,
                                            "type": f"Exam ({pn})", "body": body, "cat": cat, "url": apply_url})
                        if phase.get("result_date"):
                            events.append({"date": phase["result_date"], "exam": name,
                                            "type": f"Result ({pn})", "body": body, "cat": cat, "url": apply_url})
                except (ValueError, TypeError):
                    pass
        # Filter future events and sort
        future_events = []
        for e in events:
            try:
                d = date.fromisoformat(e["date"])
                if d >= today:
                    e["days_away"] = (d - today).days
                    future_events.append(e)
            except (ValueError, TypeError):
                pass
        future_events.sort(key=lambda x: x["date"])
        # Event type color map
        event_fills = {
            "Application Open": ACTIVE_FILL,
            "Application Close": DEADLINE_FILL,
            "Admit Card": UPCOMING_FILL,
            "Exam": PURPLE_FILL,
            "Result": TEAL_FILL,
            "Final Result": TEAL_FILL,
        }
        for idx, evt in enumerate(future_events[:500], 1):
            row = idx + 1
            values = [evt["date"], evt["exam"], evt["type"], evt["body"],
                       evt["cat"], evt["url"], evt.get("days_away")]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = THIN_BORDER
                cell.alignment = WRAP_ALIGN
            # Color code event type
            event_type = evt["type"]
            for key, fill in event_fills.items():
                if key in event_type:
                    ws.cell(row=row, column=3).fill = fill
                    break
            # Days away urgency
            self._apply_deadline_urgency(ws.cell(row=row, column=7), evt.get("days_away"))
        ws.freeze_panes = "B2"
        widths = {"A": 12, "B": 40, "C": 20, "D": 18, "E": 16, "F": 35, "G": 10}
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w
    def _build_exam_deadlines(self, wb: Workbook) -> None:
        """Sheet 13: Exam Deadlines â€” application close + admit card + result dates."""
        ws = wb.create_sheet("Exam Deadlines")
        headers = [
            "S.No", "Exam Name", "Conducting Body", "Deadline Type",
            "Date", "Days Left", "Apply URL",
        ]
        self._apply_exam_header(ws, headers)
        import json as _json
        today = date.today()
        deadlines: list[dict] = []
        # Application deadlines (within 30 days)
        for exam in self.exam_db.get_approaching_deadlines(30):
            app_end = exam.get("application_end_date")
            if app_end:
                try:
                    days = (date.fromisoformat(app_end) - today).days
                    deadlines.append({
                        "exam": exam.get("clean_exam_name") or exam.get("exam_name"),
                        "body": exam.get("conducting_body"),
                        "type": "Application Close",
                        "date": app_end, "days": days,
                        "url": exam.get("apply_online_url"),
                    })
                except ValueError:
                    pass
        # Upcoming exam dates (within 30 days)
        for exam in self.exam_db.get_upcoming_exams(30):
            phases_raw = exam.get("phases_json")
            if phases_raw:
                try:
                    for phase in _json.loads(phases_raw):
                        ed = phase.get("exam_date_start")
                        if ed:
                            try:
                                days = (date.fromisoformat(ed) - today).days
                                if 0 <= days <= 30:
                                    deadlines.append({
                                        "exam": exam.get("clean_exam_name") or exam.get("exam_name"),
                                        "body": exam.get("conducting_body"),
                                        "type": f"Exam ({phase.get('phase_name', '')})",
                                        "date": ed, "days": days,
                                        "url": exam.get("apply_online_url"),
                                    })
                            except ValueError:
                                pass
                except (ValueError, TypeError):
                    pass
        deadlines.sort(key=lambda x: x.get("days", 999))
        for idx, dl in enumerate(deadlines, 1):
            row = idx + 1
            values = [idx, dl["exam"], dl["body"], dl["type"],
                       dl["date"], dl["days"], dl["url"]]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = THIN_BORDER
                cell.alignment = WRAP_ALIGN
            self._apply_deadline_urgency(ws.cell(row=row, column=6), dl["days"])
        ws.freeze_panes = "C2"
        widths = {"A": 6, "B": 40, "C": 20, "D": 18, "E": 12, "F": 10, "G": 35}
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w
    def _build_exam_category_summary(self, wb: Workbook) -> None:
        """Sheet 14: Exam Category Summary â€” count by category."""
        ws = wb.create_sheet("Exam Category Summary")
        headers = ["Category", "Count", "% of Total"]
        self._apply_exam_header(ws, headers)
        stats = self.exam_db.get_stats()
        by_category = stats.get("by_category", {})
        total = sum(by_category.values()) or 1
        sorted_cats = sorted(by_category.items(), key=lambda x: -x[1])
        for idx, (cat, count) in enumerate(sorted_cats, 1):
            row = idx + 1
            ws.cell(row=row, column=1, value=cat.replace("_", " ")).border = THIN_BORDER
            ws.cell(row=row, column=2, value=count).border = THIN_BORDER
            pct_cell = ws.cell(row=row, column=3)
            pct_cell.value = count / total
            pct_cell.number_format = '0.0%'
            pct_cell.border = THIN_BORDER
        # Total row
        total_row = len(sorted_cats) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row - 1})").font = Font(bold=True)
        # Bar chart
        if sorted_cats:
            chart = BarChart()
            chart.type = "bar"
            chart.title = "Exams by Category"
            chart.y_axis.title = "Count"
            chart.style = 10
            data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(sorted_cats) + 1)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(sorted_cats) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.shape = 4
            chart.width = 20
            chart.height = 14
            ws.add_chart(chart, "E2")
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 12
    def _build_exam_state_summary(self, wb: Workbook) -> None:
        """Sheet 15: State Exam Summary â€” count by state for state-level exams."""
        ws = wb.create_sheet("State Exam Summary")
        headers = ["State", "Count", "% of State Total"]
        self._apply_exam_header(ws, headers)
        all_exams = self.exam_db.get_all_exams()
        by_state: dict[str, int] = {}
        for exam in all_exams:
            state = exam.get("state")
            if state:
                by_state[state] = by_state.get(state, 0) + 1
        total = sum(by_state.values()) or 1
        sorted_states = sorted(by_state.items(), key=lambda x: -x[1])
        for idx, (state, count) in enumerate(sorted_states, 1):
            row = idx + 1
            ws.cell(row=row, column=1, value=state.replace("_", " ")).border = THIN_BORDER
            ws.cell(row=row, column=2, value=count).border = THIN_BORDER
            pct_cell = ws.cell(row=row, column=3)
            pct_cell.value = count / total
            pct_cell.number_format = '0.0%'
            pct_cell.border = THIN_BORDER
        total_row = len(sorted_states) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row - 1})").font = Font(bold=True)
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 14
